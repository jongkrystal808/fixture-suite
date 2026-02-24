"""
使用記錄 API (v4.0, Stored Procedure 版)
--------------------------------------
此版本不再直接寫入 usage_logs，
所有 INSERT 一律透過 MySQL Stored Procedure：

- sp_insert_usage_log_v6

summary 表（fixture_usage_summary, serial_usage_summary）
由 Stored Procedure + 刪除時的重算邏輯共同維護。

支援：
- record_level = fixture / individual / batch
- individual / batch 會展開為 serial 紀錄呼叫 SP
- use_count > 0
- operator 若未提供 → 預設為登入者
- fixture_id / model_id / station_id 必填
"""

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from typing import Optional, Dict, Any
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from fastapi.responses import StreamingResponse
import io
from openpyxl.comments import Comment

from backend.app.database import db
from backend.app.dependencies import get_current_user, get_current_customer_id
from backend.app.utils.serial_tools import expand_serial_range, normalise_serial_list


router = APIRouter(prefix="/usage", tags=["使用紀錄 Usage Logs v4.0"])


# -------------------------------------------------------------
# Helper：基本存在性檢查（升級版：回傳 lifecycle_mode）
# -------------------------------------------------------------
def ensure_fixture_exists(fixture_id: str, customer_id: str) -> dict:
    """
    確認治具存在，並回傳其 lifecycle 設定。
    """

    rows = db.execute_query(
        """
        SELECT id, lifecycle_mode, cycle_unit
        FROM fixtures
        WHERE id=%s AND customer_id=%s
        """,
        (fixture_id, customer_id)
    )

    if not rows:
        raise HTTPException(
            status_code=400,
            detail=f"治具 {fixture_id} 不存在或不屬於客戶 {customer_id}"
        )

    return rows[0]


def ensure_model_exists(model_id: str, customer_id: str):
    row = db.execute_query(
        "SELECT id FROM machine_models WHERE id=%s AND customer_id=%s",
        (model_id, customer_id)
    )
    if not row:
        raise HTTPException(400, f"機種 {model_id} 不存在或不屬於客戶 {customer_id}")


def ensure_station_exists(station_id: str, customer_id: str):
    row = db.execute_query(
        "SELECT id FROM stations WHERE id=%s AND customer_id=%s",
        (station_id, customer_id)
    )
    if not row:
        raise HTTPException(400, f"站點 {station_id} 不存在或不屬於客戶 {customer_id}")

# -------------------------------------------------------------
# Helper：呼叫 sp_insert_usage_log_v6（✅ 把 MySQL SIGNAL / SP error 轉成 400）
# -------------------------------------------------------------
def call_sp_insert_usage(
    customer_id: str,
    fixture_id: str,
    sp_record_level: str,      # 'fixture' or 'serial'
    serial_number: Optional[str],
    station_id: str,
    model_id: str,
    use_count: int,
    operator: str,
    note: Optional[str],
):
    """
    封裝對 sp_insert_usage_log_v6 的呼叫
    並把 DB 層錯誤（SIGNAL / procedure not found / constraint）轉成 HTTP 400
    """

    try:
        out = db.call_sp_with_out(
            "sp_insert_usage_log_v6",
            [
                customer_id,
                fixture_id,
                sp_record_level,
                serial_number,
                station_id,
                model_id,
                use_count,
                operator,
                note,
            ],
            ["o_inserted_count", "o_message"],
        )
    except Exception as e:
        # pymysql 的錯誤通常在 e.args 裡，例：
        # (1644, 'serial not available: in_stock/deployed')
        # (1305, 'PROCEDURE xxx does not exist')
        msg = None
        try:
            if hasattr(e, "args") and e.args:
                # e.args 可能是 (code, message)
                if len(e.args) >= 2 and isinstance(e.args[1], str):
                    msg = e.args[1]
                else:
                    msg = str(e.args[0])
        except Exception:
            msg = None

        detail = msg or str(e) or "資料庫錯誤"

        # 常見 DB 錯誤都回 400，避免前端只看到 500
        raise HTTPException(status_code=400, detail=detail)

    # 可能是 dict 或 tuple，兩種都處理一下
    if isinstance(out, dict):
        inserted = out.get("o_inserted_count")
        message = out.get("o_message")
    else:
        inserted = out[0] if out and len(out) > 0 else None
        message = out[1] if out and len(out) > 1 else None

    if not inserted:
        raise HTTPException(400, message or "使用記錄新增失敗")

    return {
        "inserted_count": inserted,
        "message": message or "使用記錄新增成功",
    }



# -------------------------------------------------------------
# Autocomplete：搜尋機種
# -------------------------------------------------------------
@router.get("/search/models", summary="搜尋機種 (Autocomplete)")
def search_models(
    q: str,
    user=Depends(get_current_user),
    customer_id=Depends(get_current_customer_id),
):
    if not q:
        return []

    rows = db.execute_query(
        """
        SELECT id, model_name
        FROM machine_models
        WHERE customer_id=%s
          AND (id LIKE %s OR model_name LIKE %s)
        ORDER BY id
        LIMIT 20
        """,
        (customer_id, f"%{q}%", f"%{q}%"),
    )

    return rows


# -------------------------------------------------------------
# Autocomplete：搜尋站點（限定機種）
# -------------------------------------------------------------
@router.get("/search/stations", summary="搜尋站點 (依機種)")
def search_stations(
    q: str,
    model_id: str,
    user=Depends(get_current_user),
    customer_id=Depends(get_current_customer_id),
):
    if not model_id:
        raise HTTPException(400, "缺少 model_id")

    ensure_model_exists(model_id, customer_id)

    rows = db.execute_query(
        """
        SELECT s.id, s.station_name
        FROM model_stations ms
        JOIN stations s
            ON ms.station_id = s.id
            AND ms.customer_id = s.customer_id
        WHERE ms.customer_id=%s
          AND ms.model_id=%s
          AND (s.id LIKE %s OR s.station_name LIKE %s)
        ORDER BY s.id
        LIMIT 20
        """,
        (customer_id, model_id, f"%{q}%", f"%{q}%"),
    )

    return rows


# -------------------------------------------------------------
# Autocomplete：搜尋治具
# -------------------------------------------------------------
@router.get("/search/fixtures", summary="搜尋治具 (Autocomplete)")
def search_fixtures(
    q: str,
    user=Depends(get_current_user),
    customer_id=Depends(get_current_customer_id),
):
    if not q:
        return []

    rows = db.execute_query(
        """
        SELECT id, fixture_name, lifecycle_mode, cycle_unit
        FROM fixtures
        WHERE customer_id=%s
          AND (id LIKE %s OR fixture_name LIKE %s)
        ORDER BY id
        LIMIT 20
        """,
        (customer_id, f"%{q}%", f"%{q}%"),
    )

    return rows
# -------------------------------------------------------------
# 列表（僅查 usage_logs，不牽涉 summary）
# -------------------------------------------------------------
@router.get("", summary="查詢使用紀錄列表 (v4.0)")
def list_usage_logs(
    fixture_id: Optional[str] = None,
    serial_number: Optional[str] = None,
    model_id: Optional[str] = None,
    station_id: Optional[str] = None,
    operator: Optional[str] = None,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    skip: int = 0,
    limit: int = 100,
    user=Depends(get_current_user),
    customer_id=Depends(get_current_customer_id),
):
    where = ["ul.customer_id=%s"]
    params = [customer_id]

    if fixture_id:
        where.append("ul.fixture_id=%s")
        params.append(fixture_id)

    if serial_number:
        where.append("ul.serial_number=%s")
        params.append(serial_number)

    if model_id:
        where.append("ul.model_id=%s")
        params.append(model_id)

    if station_id:
        where.append("ul.station_id=%s")
        params.append(station_id)

    if operator:
        where.append("ul.operator LIKE %s")
        params.append(f"%{operator}%")

    if date_from:
        where.append("ul.used_at >= %s")
        params.append(date_from)

    if date_to:
        where.append("ul.used_at <= %s")
        params.append(date_to)

    sql = f"""
        SELECT 
            ul.*,
            f.fixture_name,
            m.model_name,
            s.station_name
        FROM usage_logs ul
        LEFT JOIN fixtures f 
            ON ul.fixture_id = f.id AND ul.customer_id = f.customer_id
        LEFT JOIN machine_models m
            ON ul.model_id = m.id AND ul.customer_id = m.customer_id
        LEFT JOIN stations s
            ON ul.station_id = s.id AND ul.customer_id = s.customer_id
        WHERE {' AND '.join(where)}
        ORDER BY ul.used_at DESC, ul.id DESC
        LIMIT %s OFFSET %s
    """

    params.extend([limit, skip])
    rows = db.execute_query(sql, tuple(params))
    return rows


# -------------------------------------------------------------
# 使用記錄匯入樣本（Template, xlsx）
# -------------------------------------------------------------
@router.get("/template", summary="下載使用記錄匯入樣本（xlsx）")
def download_usage_template(
    user=Depends(get_current_user),
    customer_id=Depends(get_current_customer_id),
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Usage Import Template"

    headers = [
        "fixture_id",
        "model_id",
        "station_id",
        "record_level",
        "serial_number",
        "serial_start",
        "serial_end",
        "use_count",
        "operator",
        "note",
    ]

    ws.append(headers)

    # 第二列：欄位說明（比照收 / 退料樣本精神）
    descriptions = [
        "治具編號（必填）",
        "機種 ID（必填）",
        "站點 ID（必填）",
        "individual / batch",
        "序號（individual 模式使用）",
        "起始序號（batch 模式使用）",
        "結束序號（batch 模式使用）",
        "使用次數（整數 > 0）",
        "操作人員（可留空，自動帶登入者）",
        "備註（選填）",
    ]
    ws.append(descriptions)

    # -----------------------------
    # 範例資料（可直接刪除）
    # -----------------------------

    # 範例 1：individual serial
    ws.append([
        "L-00062",          # fixture_id
        "AWK-1137C",        # model_id
        "T1_MAC",           # station_id
        "individual",       # record_level
        "SN0001",           # serial_number
        "",                 # serial_start
        "",                 # serial_end
        1,                  # use_count
        "",                 # operator
        "範例：單一序號使用",
    ])

    # 範例 1：batch serial
    ws.append([
        "L-00062",          # fixture_id
        "AWK-1137C",        # model_id
        "T1_MAC",           # station_id
        "batch",            # record_level
        "",                 # serial_number
        "SN0001",           # serial_start
        "SN0010",           # serial_end
        1,                  # use_count
        "",                 # operator
        "範例：批量序號使用",
    ])

    # 加上欄位註解（滑鼠 hover 可看）
    comments = {
        "D": "individual：單一序號\nbatch：序號區間",
        "E": "record_level=individual 時填",
        "F": "record_level=batch 時填",
        "G": "record_level=batch 時填",
        "H": "必須為正整數",
    }

    for col, text in comments.items():
        ws[f"{col}1"].comment = Comment(text, "system")

    # 基本欄寬
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    # 輸出
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = "usage_import_template.xlsx"

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        },
    )


# -------------------------------------------------------------
# 匯入使用記錄（Import, xlsx）
# -------------------------------------------------------------
@router.post("/import", summary="匯入使用記錄（xlsx）")
def import_usage_logs(
    file: UploadFile = File(...),
    user=Depends(get_current_user),
    customer_id=Depends(get_current_customer_id),
):
    # -----------------------------
    # 基本檢查
    # -----------------------------
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "只支援 xlsx 檔案")

    try:
        # ⭐ FastAPI UploadFile → bytes → BytesIO（最穩定）
        content = file.file.read()
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "無法讀取 Excel 檔案")

    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        raise HTTPException(400, "Excel 內容不足（至少需要標題列與資料列）")

    headers = [str(h).strip() if h else "" for h in rows[0]]

    def col(name: str):
        if name not in headers:
            raise HTTPException(400, f"缺少欄位：{name}")
        return headers.index(name)

    # 欄位索引
    idx_fixture = col("fixture_id")
    idx_model   = col("model_id")
    idx_station = col("station_id")
    idx_level   = col("record_level")
    idx_serial  = col("serial_number")
    idx_start   = col("serial_start")
    idx_end     = col("serial_end")
    idx_count   = col("use_count")
    idx_operator= col("operator")
    idx_note    = col("note")

    success = 0
    errors = []

    # -----------------------------
    # 逐列處理（從第 3 列開始）
    # -----------------------------
    for i, row in enumerate(rows[2:], start=3):
        try:
            fixture_id = str(row[idx_fixture]).strip() if row[idx_fixture] else ""
            model_id   = str(row[idx_model]).strip() if row[idx_model] else ""
            station_id = str(row[idx_station]).strip() if row[idx_station] else ""
            record_level = str(row[idx_level]).strip() if row[idx_level] else "individual"

            if not fixture_id or not model_id or not station_id:
                raise Exception("fixture_id / model_id / station_id 為必填")

            # 🔒 FINAL：fixture(datecode) 禁止寫 usage
            fixture_row = ensure_fixture_exists(fixture_id, customer_id)
            lifecycle_mode = fixture_row.get("lifecycle_mode")
            cycle_unit = fixture_row.get("cycle_unit")

            if lifecycle_mode == "fixture":
                raise Exception("Datecode 模式不可匯入 usage，請使用 days 壽命模型")

            # 🔒 FINAL：uses 僅允許 serial（individual/batch）
            if cycle_unit == "uses" and record_level not in ("individual", "batch"):
                raise Exception("uses 模式僅允許 serial 層級（individual/batch）")

            try:
                use_count = int(row[idx_count])
            except Exception:
                raise Exception("use_count 必須為整數")

            if use_count <= 0:
                raise Exception("use_count 必須 > 0")

            operator = (
                str(row[idx_operator]).strip()
                if row[idx_operator]
                else user["username"]
            )
            note = str(row[idx_note]).strip() if row[idx_note] else None


            # -----------------------------
            # individual
            # -----------------------------
            if record_level == "individual":
                sn = str(row[idx_serial]).strip() if row[idx_serial] else ""
                if not sn:
                    raise Exception("individual 模式需填 serial_number")

                call_sp_insert_usage(
                    customer_id=customer_id,
                    fixture_id=fixture_id,
                    sp_record_level="serial",
                    serial_number=sn,
                    station_id=station_id,
                    model_id=model_id,
                    use_count=use_count,
                    operator=operator,
                    note=note,
                )
                success += 1
                continue

            # -----------------------------
            # batch
            # -----------------------------
            if record_level == "batch":
                start = str(row[idx_start]).strip() if row[idx_start] else ""
                end   = str(row[idx_end]).strip() if row[idx_end] else ""

                if not start or not end:
                    raise Exception("batch 模式需填 serial_start / serial_end")

                serials = expand_serial_range(start, end)
                if not serials:
                    raise Exception("serial_start / serial_end 無法解析")

                for sn in serials:
                    call_sp_insert_usage(
                        customer_id=customer_id,
                        fixture_id=fixture_id,
                        sp_record_level="serial",
                        serial_number=sn,
                        station_id=station_id,
                        model_id=model_id,
                        use_count=use_count,
                        operator=operator,
                        note=note,
                    )
                    success += 1
                continue

            raise Exception(f"未知的 record_level：{record_level}")

        except Exception as e:
            errors.append(f"第 {i} 列：{str(e)}")

    return {
        "success_count": success,
        "error_count": len(errors),
        "errors": errors,
    }


# -------------------------------------------------------------
# 新增使用記錄（fixture / individual / batch）
# -------------------------------------------------------------
@router.post("", summary="新增使用紀錄（支援 fixture / individual / batch）")
def create_usage(
    data: Dict[str, Any],
    user=Depends(get_current_user),
    customer_id=Depends(get_current_customer_id),
):
    """
    前端 payload（app-usage.js v4.0）預期：

    {
      "fixture_id": "L-00001",
      "model_id": "EDS-2008",
      "station_id": "T1_MP",
      "use_count": 5,
      "operator": "100182",        # 可省略 → 後端帶登入者
      "note": "備註",
      "record_level": "fixture" | "individual" | "batch",
      "serials": ["SN001", "SN002"],        # individual
      "serial_start": "SN001",             # batch
      "serial_end": "SN100"                # batch
    }
    """

    fixture_id = data.get("fixture_id")
    if not fixture_id:
        raise HTTPException(400, "缺少 fixture_id")

    model_id = data.get("model_id")
    if not model_id:
        raise HTTPException(400, "缺少 model_id")

    station_id = data.get("station_id")
    if not station_id:
        raise HTTPException(400, "缺少 station_id")

    fixture_row = ensure_fixture_exists(fixture_id, customer_id)
    lifecycle_mode = fixture_row.get("lifecycle_mode")
    cycle_unit = fixture_row.get("cycle_unit")

    ensure_model_exists(model_id, customer_id)
    ensure_station_exists(station_id, customer_id)

    # 操作者
    operator = data.get("operator") or user["username"]

    # 使用次數
    try:
        use_count = int(data.get("use_count", 1))
    except Exception:
        raise HTTPException(400, "use_count 必須為整數")

    if use_count <= 0:
        raise HTTPException(400, "use_count 必須 > 0")

    note = data.get("note")

    record_level = data.get("record_level") or "individual"

    # ---------------------------------------------------------
    # 🔒 lifecycle 安全鎖定（260224 FINAL）
    # ---------------------------------------------------------
    # 1️⃣ fixture（datecode）禁止寫 usage
    if lifecycle_mode == "fixture":
        raise HTTPException(
            status_code=400,
            detail="Datecode 模式不可寫入 usage，請使用 days 壽命模型"
        )

    # 2️⃣ uses 僅允許 serial（individual/batch）
    if cycle_unit == "uses" and record_level not in ("individual", "batch"):
        raise HTTPException(
            status_code=400,
            detail="uses 模式僅允許 serial 層級使用（individual/batch）"
        )

    # -----------------------------
    # 1. individual: 多筆序號
    # -----------------------------
    if record_level == "individual":
        serials_raw = data.get("serials") or data.get("serial_numbers")
        if not serials_raw:
            raise HTTPException(400, "individual 模式需要 serials 陣列或字串")

        if isinstance(serials_raw, str):
            serials = normalise_serial_list(
                [x.strip() for x in serials_raw.split(",") if x.strip()]
            )
        else:
            serials = normalise_serial_list(serials_raw)

        if not serials:
            raise HTTPException(400, "individual 模式解析不到任何序號")

        created = 0
        errors = []

        for sn in serials:
            try:
                res = call_sp_insert_usage(
                    customer_id=customer_id,
                    fixture_id=fixture_id,
                    sp_record_level="serial",
                    serial_number=sn,
                    station_id=station_id,
                    model_id=model_id,
                    use_count=use_count,
                    operator=operator,
                    note=note,
                )
                created += res["inserted_count"] or 0
            except HTTPException as he:
                errors.append(f"{sn}: {he.detail}")
            except Exception as e:
                errors.append(f"{sn}: {str(e)}")

        return {
            "mode": "individual",
            "requested": len(serials),
            "inserted_count": created,
            "errors": errors,
        }

    # -----------------------------
    # 2. batch: serial_start + serial_end
    # -----------------------------
    if record_level == "batch":
        start = data.get("serial_start")
        end = data.get("serial_end")
        serials = expand_serial_range(start or "", end or "")

        if not serials:
            raise HTTPException(400, "batch 模式需要有效的 serial_start / serial_end")

        created = 0
        errors = []

        for sn in serials:
            try:
                res = call_sp_insert_usage(
                    customer_id=customer_id,
                    fixture_id=fixture_id,
                    sp_record_level="serial",
                    serial_number=sn,
                    station_id=station_id,
                    model_id=model_id,
                    use_count=use_count,
                    operator=operator,
                    note=note,
                )
                created += res["inserted_count"] or 0
            except HTTPException as he:
                errors.append(f"{sn}: {he.detail}")
            except Exception as e:
                errors.append(f"{sn}: {str(e)}")

        return {
            "mode": "batch",
            "range": {"start": start, "end": end},
            "requested": len(serials),
            "inserted_count": created,
            "errors": errors,
        }

    raise HTTPException(
        status_code=400,
        detail="僅支援 individual 或 batch 模式"
    )


# -------------------------------------------------------------
# 單筆查詢
# -------------------------------------------------------------
@router.get("/{log_id}", summary="取得單筆使用紀錄")
def get_usage(
    log_id: int,
    user=Depends(get_current_user),
    customer_id=Depends(get_current_customer_id),
):
    rows = db.execute_query(
        """
        SELECT 
            ul.*,
            f.fixture_name,
            m.model_name,
            s.station_name
        FROM usage_logs ul
        LEFT JOIN fixtures f 
            ON ul.fixture_id = f.id AND ul.customer_id = f.customer_id
        LEFT JOIN machine_models m
            ON ul.model_id = m.id AND ul.customer_id = m.customer_id
        LEFT JOIN stations s
            ON ul.station_id = s.id AND ul.customer_id = s.customer_id
        WHERE ul.id=%s AND ul.customer_id=%s
        """,
        (log_id, customer_id),
    )

    if not rows:
        raise HTTPException(404, "使用記錄不存在")

    return rows[0]


# -------------------------------------------------------------
# 刪除 + 自動重算 summary
# -------------------------------------------------------------
@router.delete("/{log_id}", summary="刪除使用紀錄（並重算 summary）")
def delete_usage(
    log_id: int,
    user=Depends(get_current_user),
    customer_id=Depends(get_current_customer_id),
):
    # 1. 先查舊資料（保留，方便你未來要做刪除前權限/狀態檢查）
    rows = db.execute_query(
        """
        SELECT id, customer_id, fixture_id, serial_number, record_level
        FROM usage_logs
        WHERE id=%s AND customer_id=%s
        """,
        (log_id, customer_id),
    )

    if not rows:
        raise HTTPException(404, "使用記錄不存在")

    # 2. 刪除 usage_logs（走 SP，SP 內會處理 summary 重算）
    out = db.call_sp_with_out(
        "sp_delete_usage_log_v6",
        [log_id, customer_id],
        ["o_message"]
    )

    return {
        "message": out.get("o_message") if isinstance(out, dict) else "已刪除"
    }


# -------------------------------------------------------------
# 總結摘要：治具層級
# -------------------------------------------------------------
@router.get("/summary/{fixture_id}", summary="取得治具使用摘要")
def get_fixture_summary(
    fixture_id: str,
    user=Depends(get_current_user),
    customer_id=Depends(get_current_customer_id),
):
    rows = db.execute_query(
        """
        SELECT *
        FROM fixture_usage_summary
        WHERE customer_id=%s AND fixture_id=%s
        """,
        (customer_id, fixture_id),
    )

    if not rows:
        # 沒有任何使用記錄時，回傳空物件更友善
        return {
            "customer_id": customer_id,
            "fixture_id": fixture_id,
            "total_use_count": 0,
            "last_used_at": None,
            "updated_at": None,
        }

    return rows[0]


# -------------------------------------------------------------
# 總結摘要：序號層級
# -------------------------------------------------------------
@router.get("/serial/{serial_number}", summary="取得序號使用摘要")
def get_serial_summary(
    serial_number: str,
    user=Depends(get_current_user),
    customer_id=Depends(get_current_customer_id),
):
    rows = db.execute_query(
        """
        SELECT *
        FROM serial_usage_summary
        WHERE customer_id=%s AND serial_number=%s
        """,
        (customer_id, serial_number),
    )

    if not rows:
        return {
            "customer_id": customer_id,
            "serial_number": serial_number,
            "total_use_count": 0,
            "last_used_at": None,
            "updated_at": None,
        }

    return rows[0]

# -------------------------------------------------------------
# Preview：依機種展開治具使用（支援 all / single）
# -------------------------------------------------------------
@router.post("/preview-by-model", summary="依機種加載治具使用情況（不寫入）")
def preview_by_model(
    data: Dict[str, Any],
    user=Depends(get_current_user),
    customer_id=Depends(get_current_customer_id),
):
    """
    Payload:

    全站模式:
    {
        "model_id": "EDS-2008",
        "mode": "all",
        "output_qty": 100
    }

    單站模式:
    {
        "model_id": "EDS-2008",
        "mode": "single",
        "station_id": "T1_MP",
        "run_count": 20
    }
    """

    model_id = data.get("model_id")
    mode = data.get("mode")

    if not model_id:
        raise HTTPException(400, "缺少 model_id")

    if mode not in ("all", "single"):
        raise HTTPException(400, "mode 必須為 all 或 single")

    ensure_model_exists(model_id, customer_id)

    # ---------------------------------------------------------
    # 取得 model 底下站點
    # ---------------------------------------------------------
    station_rows = db.execute_query(
        """
        SELECT station_id
        FROM model_stations
        WHERE customer_id=%s AND model_id=%s
        """,
        (customer_id, model_id),
    )

    if not station_rows:
        return {
            "model_id": model_id,
            "stations": [],
        }

    # ---------------------------------------------------------
    # 決定要展開哪些站點
    # ---------------------------------------------------------
    if mode == "all":
        output_qty = data.get("output_qty")

        try:
            output_qty = int(output_qty)
        except Exception:
            raise HTTPException(400, "output_qty 必須為整數")

        if output_qty <= 0:
            raise HTTPException(400, "output_qty 必須 > 0")

        selected_stations = [s["station_id"] for s in station_rows]

    else:  # single
        station_id = data.get("station_id")
        run_count = data.get("run_count")

        if not station_id:
            raise HTTPException(400, "single 模式需提供 station_id")

        if station_id not in [s["station_id"] for s in station_rows]:
            raise HTTPException(400, "該站點不屬於此機種")

        try:
            run_count = int(run_count)
        except Exception:
            raise HTTPException(400, "run_count 必須為整數")

        if run_count <= 0:
            raise HTTPException(400, "run_count 必須 > 0")

        selected_stations = [station_id]

    # ---------------------------------------------------------
    # 查每個站點的治具需求
    # ---------------------------------------------------------
    preview_result = []

    for station_id in selected_stations:

        fixture_rows = db.execute_query(
            """
            SELECT fr.fixture_id,
                   fr.required_qty,
                   f.fixture_name,
                   f.lifecycle_mode,
                   f.cycle_unit
            FROM fixture_requirements fr
                     JOIN fixtures f
                          ON fr.fixture_id = f.id
                              AND fr.customer_id = f.customer_id
            WHERE fr.customer_id = %s
              AND fr.model_id = %s
              AND fr.station_id = %s
            """,
            (customer_id, model_id, station_id),
        )

        station_result = []

        for r in fixture_rows:

            lifecycle_mode = r["lifecycle_mode"]
            cycle_unit = r["cycle_unit"]
            required_qty = r["required_qty"]

            # 🔥 不再 skip days
            is_datecode = lifecycle_mode == "fixture"

            # uses 才會真正寫 usage
            is_readonly = (cycle_unit != "uses") or is_datecode

            if mode == "all":
                use_count = output_qty * required_qty
                base_value = output_qty
                base_label = "output_qty"
            else:
                use_count = run_count * required_qty
                base_value = run_count
                base_label = "run_count"

            station_result.append({
                "station_id": station_id,
                "fixture_id": r["fixture_id"],
                "fixture_name": r["fixture_name"],
                "required_qty": required_qty,
                "use_count": use_count,
                "readonly": is_readonly,
                "lifecycle_mode": lifecycle_mode,
                "cycle_unit": cycle_unit,
                "formula": f"{base_label}({base_value}) × required_qty({required_qty}) = {use_count}"
            })

        if station_result:
            preview_result.append({
                "station_id": station_id,
                "fixtures": station_result
            })

    return {
        "model_id": model_id,
        "mode": mode,
        "stations": preview_result
    }

# -------------------------------------------------------------
# Create：依機種批量寫入 usage（每個序號各記完整 use_count）
# -------------------------------------------------------------
@router.post("/create-by-model", summary="依機種寫入使用紀錄")
def create_by_model(
    data: Dict[str, Any],
    user=Depends(get_current_user),
    customer_id=Depends(get_current_customer_id),
):
    """
    Payload 範例:

    {
        "model_id": "EDS-2008",
        "mode": "all",
        "operator": "100182",
        "records": [
            {
                "station_id": "T1_MP",
                "fixture_id": "L-0001",
                "use_count": 200,
                "serials": "SN001,SN002",
                "formula": "output_qty(100) × required_qty(2) = 200"
            }
        ]
    }
    """

    model_id = data.get("model_id")
    mode = data.get("mode")
    operator = data.get("operator") or user["username"]
    records = data.get("records")

    if not model_id:
        raise HTTPException(400, "缺少 model_id")

    if mode not in ("all", "single"):
        raise HTTPException(400, "mode 必須為 all 或 single")

    if not records or not isinstance(records, list):
        raise HTTPException(400, "records 必須為陣列")

    ensure_model_exists(model_id, customer_id)

    total_inserted = 0
    errors = []

    for item in records:

        station_id = item.get("station_id")
        fixture_id = item.get("fixture_id")
        use_count = item.get("use_count")
        serials_raw = item.get("serials")
        formula = item.get("formula", "")

        if not station_id or not fixture_id:
            errors.append(f"{fixture_id}: station_id 或 fixture_id 缺失")
            continue

        try:
            use_count = int(use_count)
        except Exception:
            errors.append(f"{fixture_id}: use_count 必須為整數")
            continue

        if use_count <= 0:
            errors.append(f"{fixture_id}: use_count 必須 > 0")
            continue

        # 確認站點存在
        ensure_station_exists(station_id, customer_id)

        # 查治具資訊
        fixture_row = ensure_fixture_exists(fixture_id, customer_id)

        lifecycle_mode = fixture_row.get("lifecycle_mode")
        cycle_unit = fixture_row.get("cycle_unit")

        # 🔒 跳過 datecode
        if lifecycle_mode == "fixture":
            continue

        # 🔒 只允許 serial + uses
        if cycle_unit != "uses":
            continue

        if not serials_raw:
            errors.append(f"{fixture_id}: 未填寫序號")
            continue

        if isinstance(serials_raw, str):
            serial_list = normalise_serial_list(
                [x.strip() for x in serials_raw.split(",") if x.strip()]
            )
        else:
            serial_list = normalise_serial_list(serials_raw)

        if not serial_list:
            errors.append(f"{fixture_id}: 無有效序號")
            continue

        # 每個序號各記完整 use_count
        for sn in serial_list:
            try:
                result = call_sp_insert_usage(
                    customer_id=customer_id,
                    fixture_id=fixture_id,
                    sp_record_level="serial",
                    serial_number=sn,
                    station_id=station_id,
                    model_id=model_id,
                    use_count=use_count,
                    operator=operator,
                    note=formula,
                )
                total_inserted += result["inserted_count"] or 0
            except HTTPException as he:
                errors.append(f"{fixture_id}-{sn}: {he.detail}")
            except Exception as e:
                errors.append(f"{fixture_id}-{sn}: {str(e)}")

    return {
        "model_id": model_id,
        "mode": mode,
        "total_inserted": total_inserted,
        "error_count": len(errors),
        "errors": errors,
    }