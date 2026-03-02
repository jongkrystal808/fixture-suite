"""
更換記錄 API (v6 FINAL)
------------------------------------------------
完全對齊 sp_fixture_replacement_master_v3

規則：
- record_level = serial / fixture
- fixture 模式：scrap / maintenance 都必須 scrap_qty
- lifecycle_mode=fixture（datecode）→ 必須提供 datecode
- datecode 透過 serial_number 傳入 SP
"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from typing import Optional, Dict, Any
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from fastapi.responses import StreamingResponse
import io

from backend.app.database import db
from backend.app.dependencies import get_current_user, get_current_customer_id

router = APIRouter(
    prefix="/replacement",
    tags=["更換紀錄 Replacement Logs v6"]
)

# =============================================================
# Helper
# =============================================================

def ensure_fixture_exists(fixture_id: str, customer_id: str) -> dict:
    rows = db.execute_query(
        """
        SELECT id, lifecycle_mode, cycle_unit
        FROM fixtures
        WHERE id=%s AND customer_id=%s
        """,
        (fixture_id, customer_id)
    )

    if not rows:
        raise HTTPException(
            status_code=400,
            detail=f"治具 {fixture_id} 不存在或不屬於客戶 {customer_id}"
        )

    return rows[0]


def decorate_note(event_type: str, note: Optional[str]) -> Optional[str]:
    base = (note or "").strip()
    prefix = f"[{event_type}]"
    return f"{prefix} {base}" if base else prefix


def call_sp_fixture_replacement_master_v3(
    customer_id: str,
    fixture_id: str,
    record_level: str,
    event_type: str,
    serial_number: Optional[str],
    scrap_qty: Optional[int],
    operator: str,
    note: Optional[str],
) -> Dict[str, Any]:

    occurred_at = datetime.now()

    try:
        out = db.call_sp_with_out(
            "sp_fixture_replacement_master_v3",
            [
                customer_id,
                fixture_id,
                record_level,
                event_type,
                serial_number,
                scrap_qty,
                operator,
                note,
                occurred_at,
            ],
            ["o_inserted_count", "o_message"],
        )

    except Exception as e:
        msg = None
        if hasattr(e, "args") and e.args:
            msg = e.args[1] if len(e.args) >= 2 else str(e.args[0])
        raise HTTPException(status_code=400, detail=msg or str(e))

    if isinstance(out, dict):
        inserted = out.get("o_inserted_count")
        message = out.get("o_message")
    else:
        inserted = out[0] if out else None
        message = out[1] if out and len(out) > 1 else None

    return {
        "inserted_count": inserted,
        "message": message or "Replacement 完成"
    }


# =============================================================
# 查詢列表
# =============================================================

@router.get("", summary="查詢更換紀錄列表")
def list_replacement_logs(
    fixture_id: Optional[str] = None,
    serial_number: Optional[str] = None,
    record_level: Optional[str] = None,
    event_type: Optional[str] = None,
    operator: Optional[str] = None,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    skip: int = 0,
    limit: int = 100,
    user=Depends(get_current_user),
    customer_id=Depends(get_current_customer_id),
):
    where = ["rl.customer_id=%s"]
    params = [customer_id]

    if fixture_id:
        where.append("rl.fixture_id=%s")
        params.append(fixture_id)

    if serial_number:
        where.append("rl.serial_number=%s")
        params.append(serial_number)

    if record_level:
        where.append("rl.record_level=%s")
        params.append(record_level)

    if event_type:
        where.append("rl.event_type=%s")
        params.append(event_type)

    if operator:
        where.append("rl.operator LIKE %s")
        params.append(f"%{operator}%")

    if date_from:
        where.append("rl.occurred_at >= %s")
        params.append(date_from)

    if date_to:
        where.append("rl.occurred_at <= %s")
        params.append(date_to)

    sql = f"""
        SELECT
            rl.id,
            rl.fixture_id,
            f.fixture_name,
            rl.record_level,
            rl.event_type,
            rl.serial_number,
            rl.scrap_qty,
            rl.operator,
            rl.note,
            rl.occurred_at
        FROM replacement_logs rl
        LEFT JOIN fixtures f
          ON rl.fixture_id = f.id
         AND rl.customer_id = f.customer_id
        WHERE {' AND '.join(where)}
        ORDER BY rl.occurred_at DESC, rl.id DESC
        LIMIT %s OFFSET %s
    """

    params.extend([limit, skip])
    return db.execute_query(sql, tuple(params))
# =============================================================
# 新增更換記錄
# =============================================================

@router.post("", summary="新增更換紀錄")
def create_replacement(
    data: Dict[str, Any],
    user=Depends(get_current_user),
    customer_id=Depends(get_current_customer_id),
):

    fixture_id = data.get("fixture_id")
    if not fixture_id:
        raise HTTPException(400, "缺少 fixture_id")

    fixture_row = ensure_fixture_exists(fixture_id, customer_id)
    lifecycle_mode = fixture_row.get("lifecycle_mode")

    record_level = data.get("record_level")
    if record_level not in ("serial", "fixture"):
        raise HTTPException(400, "record_level 必須為 serial 或 fixture")

    event_type = data.get("event_type")
    if event_type not in ("scrap", "maintenance"):
        raise HTTPException(400, "event_type 必須為 scrap 或 maintenance")

    operator = data.get("operator") or user["username"]

    datecode = data.get("datecode")
    if datecode:
        datecode = str(datecode).strip()

    serial_number = None
    scrap_qty = None

    # ---------------------------------------------------------
    # serial 模式
    # ---------------------------------------------------------
    if record_level == "serial":

        serial_number = data.get("serial_number")
        if not serial_number:
            raise HTTPException(400, "serial 模式需提供 serial_number")

    # ---------------------------------------------------------
    # fixture 模式
    # ---------------------------------------------------------
    if record_level == "fixture":

        raw_qty = data.get("scrap_qty")
        if raw_qty is None:
            raise HTTPException(400, "fixture 模式必須提供 scrap_qty")

        try:
            scrap_qty = int(raw_qty)
        except Exception:
            raise HTTPException(400, "scrap_qty 必須為整數")

        if scrap_qty <= 0:
            raise HTTPException(400, "scrap_qty 必須 > 0")

        # lifecycle_mode=fixture（datecode）
        if lifecycle_mode == "fixture":
            if not datecode:
                raise HTTPException(400, "datecode 模式必須提供 datecode")
            serial_number = datecode  # 傳入 SP

    note = decorate_note(event_type, data.get("note"))

    return call_sp_fixture_replacement_master_v3(
        customer_id=customer_id,
        fixture_id=fixture_id,
        record_level=record_level,
        event_type=event_type,
        serial_number=serial_number,
        scrap_qty=scrap_qty,
        operator=operator,
        note=note,
    )


# =============================================================
# 下載匯入樣本
# =============================================================

@router.get("/template", summary="下載更換記錄匯入樣本")
def download_replacement_template(
    user=Depends(get_current_user),
    customer_id=Depends(get_current_customer_id),
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Replacement Import Template"

    headers = [
        "fixture_id",
        "record_level",
        "event_type",
        "serial_number",
        "datecode",
        "scrap_qty",
        "operator",
        "note",
    ]
    ws.append(headers)

    descriptions = [
        "治具編號（必填）",
        "serial / fixture",
        "scrap / maintenance",
        "序號（record_level=serial 必填）",
        "日期碼（lifecycle_mode=fixture 且 record_level=fixture 必填）",
        "數量（record_level=fixture 必填）",
        "操作人員（可留空）",
        "備註",
    ]
    ws.append(descriptions)

    # Serial scrap
    ws.append([
        "L-00062",
        "serial",
        "scrap",
        "SN0001",
        "",
        "",
        "",
        "單顆序號報廢",
    ])

    # Serial maintenance
    ws.append([
        "L-00062",
        "serial",
        "maintenance",
        "SN0002",
        "",
        "",
        "",
        "單顆送修",
    ])

    # Fixture scrap (datecode)
    ws.append([
        "L-00062",
        "fixture",
        "scrap",
        "",
        "20260224",
        5,
        "",
        "批次報廢 5",
    ])

    # Fixture maintenance
    ws.append([
        "L-00062",
        "fixture",
        "maintenance",
        "",
        "20260224",
        3,
        "",
        "批次送修 3",
    ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 26

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="replacement_import_template.xlsx"'
        },
    )


# =============================================================
# 匯入
# =============================================================

@router.post("/import", summary="匯入更換記錄（xlsx）")
def import_replacement_logs(
    file: UploadFile = File(...),
    user=Depends(get_current_user),
    customer_id=Depends(get_current_customer_id),
):

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "只支援 xlsx")

    try:
        content = file.file.read()
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "無法讀取 Excel")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 3:
        raise HTTPException(400, "Excel 內容不足")

    headers = [str(h).strip() if h else "" for h in rows[0]]

    def col(name):
        if name not in headers:
            raise HTTPException(400, f"缺少欄位：{name}")
        return headers.index(name)

    idx_fixture = col("fixture_id")
    idx_level = col("record_level")
    idx_event = col("event_type")
    idx_serial = col("serial_number")
    idx_datecode = col("datecode")
    idx_qty = col("scrap_qty")
    idx_operator = col("operator")
    idx_note = col("note")

    success = 0
    errors = []

    for i, row in enumerate(rows[2:], start=3):
        try:
            fixture_id = str(row[idx_fixture]).strip()
            record_level = str(row[idx_level]).strip()
            event_type = str(row[idx_event]).strip()
            serial_number = str(row[idx_serial]).strip() if row[idx_serial] else None
            datecode = str(row[idx_datecode]).strip() if row[idx_datecode] else None
            scrap_qty = row[idx_qty]

            operator = (
                str(row[idx_operator]).strip()
                if row[idx_operator]
                else user["username"]
            )

            note = str(row[idx_note]).strip() if row[idx_note] else None

            payload = {
                "fixture_id": fixture_id,
                "record_level": record_level,
                "event_type": event_type,
                "serial_number": serial_number,
                "datecode": datecode,
                "scrap_qty": scrap_qty,
                "operator": operator,
                "note": note,
            }

            create_replacement(payload, user, customer_id)

            success += 1

        except Exception as e:
            errors.append(f"第 {i} 列：{str(e)}")

    return {
        "success_count": success,
        "error_count": len(errors),
        "errors": errors,
    }

@router.get("/fixture-info")
def get_fixture_info(
    fixture_id: str,
    user=Depends(get_current_user),
    customer_id=Depends(get_current_customer_id),
):
    row = db.execute_query(
        """
        SELECT lifecycle_mode
        FROM fixtures
        WHERE id=%s AND customer_id=%s
        """,
        (fixture_id, customer_id)
    )

    if not row:
        raise HTTPException(404, "fixture not found")

    return row[0]