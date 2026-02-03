"""
機種 / 機種-站點-治具 Excel 匯入 Router（v4.x / Header-based）

行為規則：
- 若 Excel 僅包含 model 資料 → 只匯入 machine_models
- 若 Excel 包含 station / fixture 欄位 → 同時處理綁定與治具需求
"""

from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
from openpyxl import load_workbook
from io import BytesIO

from backend.app.database import db
from backend.app.dependencies import get_current_user, get_current_customer_id

router = APIRouter(
    prefix="/models",
    tags=["Machine Models Import"]
)

MODEL_SHEET = "models"


@router.post("/import", summary="匯入機種 / 站點 / 治具（XLSX）")
async def import_models_xlsx(
    file: UploadFile = File(...),
    user=Depends(get_current_user),
    customer_id: str = Depends(get_current_customer_id),
):

    # =================================================
    # 1️⃣ 基本檢查
    # =================================================
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="只接受 .xlsx 檔案")

    content = await file.read()
    wb = load_workbook(BytesIO(content))

    if MODEL_SHEET not in wb.sheetnames:
        raise HTTPException(
            status_code=400,
            detail='XLSX 需包含 "models" 工作表'
        )

    ws = wb[MODEL_SHEET]

    # =================================================
    # 2️⃣ Header 解析（自動判斷模式）
    # =================================================
    header = [str(c.value).strip() if c.value else "" for c in ws[1]]

    # 基本欄位（一定要有）
    BASE_COLS = ["id", "model_name"]
    for col in BASE_COLS:
        if col not in header:
            raise HTTPException(
                status_code=400,
                detail=f'缺少必要欄位：{col}'
            )

    # 可選欄位
    HAS_NOTE = "note" in header

    # 判斷是否為「進階模式」
    ADVANCED_COLS = ["station_id", "fixture_id", "required_qty"]
    IS_ADVANCED = all(col in header for col in ADVANCED_COLS)

    idx = {name: header.index(name) for name in header}

    # =================================================
    # 3️⃣ 初始化結果
    # =================================================
    result = {
        "models": {"imported": 0, "updated": 0, "skipped": 0},
        "errors": [],
    }

    if IS_ADVANCED:
        result["model_station_fixture"] = {
            "imported": 0,
            "updated": 0,
            "skipped": 0,
        }

    conn = db.get_conn()

    try:
        cursor = conn.cursor()

        # =================================================
        # 4️⃣ 逐行處理
        # =================================================
        for row_no, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True),
            start=2
        ):
            try:
                if not row or not row[idx["id"]] or not row[idx["model_name"]]:
                    result["models"]["skipped"] += 1
                    continue

                model_id = str(row[idx["id"]]).strip()
                model_name = str(row[idx["model_name"]]).strip()
                note = (
                    str(row[idx["note"]]).strip()
                    if HAS_NOTE and row[idx.get("note")] else ""
                )

                # -----------------------------
                # 4-1. machine_models upsert
                # -----------------------------
                cursor.execute(
                    """
                    SELECT id
                    FROM machine_models
                    WHERE customer_id=%s AND id=%s
                    """,
                    (customer_id, model_id)
                )
                exist = cursor.fetchone()

                if exist:
                    cursor.execute(
                        """
                        UPDATE machine_models
                        SET model_name=%s, note=%s
                        WHERE customer_id=%s AND id=%s
                        """,
                        (model_name, note, customer_id, model_id)
                    )
                    result["models"]["updated"] += 1
                else:
                    cursor.execute(
                        """
                        INSERT INTO machine_models
                            (customer_id, id, model_name, note)
                        VALUES (%s,%s,%s,%s)
                        """,
                        (customer_id, model_id, model_name, note)
                    )
                    result["models"]["imported"] += 1

                # -----------------------------
                # 4-2. 若非進階模式，直接下一行
                # -----------------------------
                if not IS_ADVANCED:
                    continue

                # -----------------------------
                # 4-3. 解析進階欄位
                # -----------------------------
                station_id = str(row[idx["station_id"]]).strip()
                fixture_id = str(row[idx["fixture_id"]]).strip()
                required_qty = row[idx["required_qty"]]

                if not station_id or not fixture_id:
                    result["model_station_fixture"]["skipped"] += 1
                    continue

                if required_qty is None or int(required_qty) < 0:
                    raise ValueError("required_qty 必須 >= 0")

                required_qty = int(required_qty)

                # -----------------------------
                # 確保 station / fixture 存在
                # -----------------------------
                cursor.execute(
                    "SELECT 1 FROM stations WHERE customer_id=%s AND id=%s",
                    (customer_id, station_id)
                )
                if not cursor.fetchone():
                    raise ValueError(f"站點不存在：{station_id}")

                cursor.execute(
                    "SELECT 1 FROM fixtures WHERE customer_id=%s AND id=%s",
                    (customer_id, fixture_id)
                )
                if not cursor.fetchone():
                    raise ValueError(f"治具不存在：{fixture_id}")

                # -----------------------------
                # 補 model_stations
                # -----------------------------
                cursor.execute(
                    """
                    SELECT 1 FROM model_stations
                    WHERE customer_id=%s AND model_id=%s AND station_id=%s
                    """,
                    (customer_id, model_id, station_id)
                )
                if not cursor.fetchone():
                    cursor.execute(
                        """
                        INSERT INTO model_stations
                            (customer_id, model_id, station_id)
                        VALUES (%s,%s,%s)
                        """,
                        (customer_id, model_id, station_id)
                    )

                # -----------------------------
                # fixture_requirements upsert
                # -----------------------------
                cursor.execute(
                    """
                    SELECT id FROM fixture_requirements
                    WHERE customer_id=%s
                      AND model_id=%s
                      AND station_id=%s
                      AND fixture_id=%s
                    """,
                    (customer_id, model_id, station_id, fixture_id)
                )
                fr = cursor.fetchone()

                if fr:
                    cursor.execute(
                        """
                        UPDATE fixture_requirements
                        SET required_qty=%s
                        WHERE id=%s AND customer_id=%s
                        """,
                        (required_qty, fr[0], customer_id)
                    )
                    result["model_station_fixture"]["updated"] += 1
                else:
                    cursor.execute(
                        """
                        INSERT INTO fixture_requirements
                            (customer_id, model_id, station_id, fixture_id, required_qty)
                        VALUES (%s,%s,%s,%s,%s)
                        """,
                        (customer_id, model_id, station_id, fixture_id, required_qty)
                    )
                    result["model_station_fixture"]["imported"] += 1

            except Exception as e:
                result["errors"].append(
                    f"第 {row_no} 行：{str(e)}"
                )

        conn.commit()

    finally:
        conn.close()

    return {
        "message": "匯入完成",
        **result
    }
