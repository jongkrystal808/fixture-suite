from fastapi import APIRouter, Depends, HTTPException, Query, status, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import Optional, List
import traceback
from openpyxl import Workbook, load_workbook
from io import BytesIO
from backend.app.database import db
from backend.app.dependencies import get_current_user, get_current_customer_id, get_current_admin
from backend.app.models.fixture import (
    FixtureCreate,
    FixtureUpdate,
    FixtureResponse,
    FixtureSimple,
    FixtureStatusViewRow,
    FixtureStatistics,
    CycleUnit
)
router = APIRouter(
    prefix="/fixtures",
    tags=["治具管理 Fixtures"]
)

# ============================================================
# ⭐ 重要：所有具體路徑必須在動態路徑 /{fixture_id} 之前定義
# ============================================================


# ============================================================
# 📦 治具儲位清單（一定要在動態路徑之前）
# ============================================================
@router.get(
    "/storages",
    summary="取得治具儲位清單"
)
async def list_fixture_storages(
    customer_id: str = Depends(get_current_customer_id),
):
    rows = db.execute_query(
        """
        SELECT DISTINCT storage_location
        FROM fixtures
        WHERE customer_id = %s
          AND is_scrapped = 0
          AND storage_location IS NOT NULL
          AND storage_location <> ''
        ORDER BY storage_location
        """,
        (customer_id,)
    )

    return [r["storage_location"] for r in rows]


# ============================================================
# 🔍 治具模糊搜尋
# ============================================================
@router.get("/search", summary="治具模糊搜尋")
async def search_fixtures(
    q: str = Query(..., min_length=1),
    limit: int = Query(20, le=50),
    customer_id: str = Depends(get_current_customer_id),
    user=Depends(get_current_user),
):
    try:
        rows = db.execute_query(
            """
            SELECT id AS fixture_id, fixture_name
            FROM fixtures
            WHERE customer_id = %s
              AND is_scrapped = 0
              AND LOWER(id) LIKE LOWER(%s)
            ORDER BY id
            LIMIT %s
            """,
            (customer_id, f"%{q}%", limit)
        )
        return rows or []

    except Exception:
        print("❌ [search_fixtures]\n", traceback.format_exc())
        raise HTTPException(status_code=500, detail="查詢治具失敗")


# ============================================================
# 📊 治具統計摘要
# ============================================================
@router.get(
    "/statistics/summary",
    response_model=FixtureStatistics,
    summary="治具統計摘要"
)
async def fixture_statistics(
    customer_id: str = Depends(get_current_customer_id),
    user=Depends(get_current_user),
):
    try:
        row = db.execute_query(
            """
            SELECT
                COUNT(*) AS total_fixtures,
                SUM(in_stock_qty) AS in_stock_qty,
                SUM(deployed_qty) AS deployed_qty,
                SUM(maintenance_qty) AS maintenance_qty,
                SUM(scrapped_qty) AS scrapped_qty,
                SUM(returned_qty) AS returned_qty
            FROM fixtures
            WHERE customer_id = %s
            """,
            (customer_id,)
        )[0]

        return {
            "total_fixtures": row["total_fixtures"] or 0,
            "in_stock_qty": row["in_stock_qty"] or 0,
            "deployed_qty": row["deployed_qty"] or 0,
            "maintenance_qty": row["maintenance_qty"] or 0,
            "scrapped_qty": row["scrapped_qty"] or 0,
            "returned_qty": row["returned_qty"] or 0,
        }

    except Exception:
        print("❌ [fixture_statistics]\n", traceback.format_exc())
        raise HTTPException(status_code=500, detail="查詢統計資料失敗")


# ============================================================
# 📊 治具狀態總覽（View）
# ============================================================
@router.get(
    "/status/view",
    response_model=List[FixtureStatusViewRow],
    summary="治具狀態總覽"
)
async def fixture_status_view(
    skip: int = 0,
    limit: int = 100,
    customer_id: str = Depends(get_current_customer_id),
    user=Depends(get_current_user),
):
    try:
        rows = db.execute_query(
            """
            SELECT *
            FROM view_fixture_status
            WHERE customer_id = %s
            ORDER BY fixture_id
            LIMIT %s OFFSET %s
            """,
            (customer_id, limit, skip)
        )
        return [FixtureStatusViewRow(**row) for row in rows]

    except Exception:
        print("❌ [fixture_status_view]\n", traceback.format_exc())
        raise HTTPException(status_code=500, detail="查詢治具狀態失敗")


# ============================================================
# 📋 簡化治具列表（下拉選單）
# ============================================================
@router.get(
    "/simple/list",
    response_model=List[FixtureSimple],
    summary="簡化治具列表"
)
async def simple_fixture_list(
    customer_id: str = Depends(get_current_customer_id),
    user=Depends(get_current_user),
):
    try:
        rows = db.execute_query(
            """
            SELECT
                id AS fixture_id,
                fixture_name
            FROM fixtures
            WHERE customer_id = %s
            AND is_scrapped = 0
            ORDER BY id
            """,
            (customer_id,)
        )
        return [FixtureSimple(**row) for row in rows]

    except Exception:
        print("❌ [simple_fixture_list]\n", traceback.format_exc())
        raise HTTPException(status_code=500, detail="查詢簡化治具列表失敗")


# ============================================================
# 📥 下載治具匯入樣本
# ============================================================
@router.get("/template", summary="下載治具匯入樣本（XLSX）")
async def download_fixtures_template(
    customer_id: str = Depends(get_current_customer_id),
    user=Depends(get_current_user),
):
    wb = Workbook()
    ws = wb.active
    ws.title = "fixtures"

    ws.append([
        "id",
        "fixture_name",
        "fixture_type",
        "storage_location",
        "replacement_cycle",
        "cycle_unit",
        "note",
    ])

    ws.append([
        "FX-001",
        "治具範例一",
        "ICT",
        "A01",
        1000,
        "uses",
        "備註示例",
    ])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            'attachment; filename="fixtures_import_template.xlsx"'
        }
    )



# ============================================================
# 📤 匯出治具（XLSX）- Header-based customer context
# （必須在 /{fixture_id} 之前）
# ============================================================
@router.get("/export", summary="匯出治具（XLSX）")
async def export_fixtures_xlsx(
    user=Depends(get_current_user),
    customer_id: str = Depends(get_current_customer_id),
):
    rows = db.execute_query(
        """
        SELECT
            id,
            fixture_name,
            fixture_type,
            storage_location,
            replacement_cycle,
            cycle_unit,
            note
        FROM fixtures
        WHERE customer_id = %s
        ORDER BY id
        """,
        (customer_id,)
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "fixtures"

    ws.append([
        "id",
        "fixture_name",
        "fixture_type",
        "storage_location",
        "replacement_cycle",
        "cycle_unit",
        "note",
    ])

    for r in rows:
        ws.append([
            r["id"],
            r.get("fixture_name") or "",
            r.get("fixture_type") or "",
            r.get("storage_location") or "",
            r.get("replacement_cycle") or "",
            r.get("cycle_unit") or "",
            r.get("note") or "",
        ])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            f'attachment; filename="fixtures_{customer_id}.xlsx"'
        }
    )



# ============================================================
# ⚠️ 以下是動態路徑，必須放在所有具體路徑之後
# ============================================================


# ============================================================
# 建立治具 (CREATE)
# ============================================================
@router.post(
    "",
    response_model=FixtureResponse,
    status_code=status.HTTP_201_CREATED,
    summary="建立新治具"
)
async def create_fixture(
    data: FixtureCreate,
    user=Depends(get_current_user),
    customer_id: str = Depends(get_current_customer_id),
):
    try:
        # -------------------------------------------------
        # 1️⃣ 檢查治具是否已存在（同客戶）
        # -------------------------------------------------
        exists = db.execute_query(
            """
            SELECT id
            FROM fixtures
            WHERE id = %s AND customer_id = %s
            """,
            (data.id, customer_id)
        )
        if exists:
            raise HTTPException(
                status_code=400,
                detail=f"治具 {data.id} 已存在"
            )

        # -------------------------------------------------
        # 2️⃣ 插入新治具
        # -------------------------------------------------
        sql = """
            INSERT INTO fixtures (
                id,
                customer_id,
                fixture_name,
                fixture_type,
                self_purchased_qty,
                customer_supplied_qty,
                in_stock_qty,
                deployed_qty,
                maintenance_qty,
                scrapped_qty,
                returned_qty,
                storage_location,
                replacement_cycle,
                cycle_unit,
                owner_id,
                note
            ) VALUES (
                %s, %s, %s, %s,
                %s, %s,
                %s, %s, %s, %s, %s,
                %s,
                %s, %s,
                %s,
                %s
            )
        """

        params = (
            data.id,
            customer_id,
            data.fixture_name,
            data.fixture_type,
            data.self_purchased_qty or 0,
            data.customer_supplied_qty or 0,

            # 初始化數量（v4.x 建立時統一 0）
            0,  # in_stock_qty
            0,  # deployed_qty
            0,  # maintenance_qty
            0,  # scrapped_qty
            0,  # returned_qty

            data.storage_location,
            data.replacement_cycle,
            data.cycle_unit.value if hasattr(data.cycle_unit, "value") else data.cycle_unit,
            data.owner_id,
            data.note,
        )

        db.execute_update(sql, params)

        # -------------------------------------------------
        # 3️⃣ 回傳新建立的治具
        # -------------------------------------------------
        row = db.execute_query(
            """
            SELECT *
            FROM fixtures
            WHERE id = %s AND customer_id = %s
            """,
            (data.id, customer_id)
        )[0]

        return FixtureResponse(**row)

    except HTTPException:
        raise

    except Exception:
        print("❌ [create_fixture]\n", traceback.format_exc())
        raise HTTPException(status_code=500, detail="建立治具失敗")



# ============================================================
# 取得治具詳情 (READ) - 動態路徑
# ============================================================
@router.get("/{fixture_id}", response_model=FixtureResponse, summary="取得治具詳情")
async def get_fixture(
    fixture_id: str,
    user=Depends(get_current_user),
    customer_id: str = Depends(get_current_customer_id),
):
    try:
        sql = """
            SELECT
                f.id,
                f.customer_id,
                f.fixture_name,
                f.fixture_type,
                f.self_purchased_qty,
                f.customer_supplied_qty,
                f.in_stock_qty,
                f.deployed_qty,
                f.maintenance_qty,
                f.scrapped_qty,
                f.returned_qty,
                f.storage_location,
                f.replacement_cycle,
                f.cycle_unit,
                f.last_replacement_date,
                f.last_notification_time,
                f.owner_id,
                f.note,
                f.created_at,
                f.updated_at,
                u.username AS owner_name
            FROM fixtures f
            LEFT JOIN owners o ON f.owner_id = o.id
            LEFT JOIN users u ON u.id = o.primary_user_id
            WHERE f.id = %s AND f.customer_id = %s
            LIMIT 1
        """

        rows = db.execute_query(sql, (fixture_id, customer_id))
        if not rows:
            raise HTTPException(status_code=404, detail="治具不存在")

        return FixtureResponse(**rows[0])

    except HTTPException:
        raise
    except Exception:
        print("❌ [get_fixture]\n", traceback.format_exc())
        raise HTTPException(status_code=500, detail="查詢治具失敗")


# ============================================================
# 取得治具完整詳細資料 (DETAIL) - 動態路徑
# ============================================================
@router.get("/{fixture_id}/detail", summary="取得治具完整詳細資料")
async def get_fixture_detail(
    fixture_id: str,
    user=Depends(get_current_user),
    customer_id: str = Depends(get_current_customer_id),
):
    try:
        # 1️⃣ 基本資料
        fixture_sql = """
            SELECT
                f.*,
                u.username AS owner_name
            FROM fixtures f
            LEFT JOIN owners o ON f.owner_id = o.id
            LEFT JOIN users u ON u.id = o.primary_user_id
            WHERE f.id = %s AND f.customer_id = %s
            LIMIT 1
        """
        fixture_rows = db.execute_query(fixture_sql, (fixture_id, customer_id))
        if not fixture_rows:
            raise HTTPException(status_code=404, detail="治具不存在")

        # 2️⃣ 收 / 退料
        last_receipt = db.execute_query(
            """
            SELECT id, transaction_date, order_no, operator, note
            FROM material_transactions
            WHERE fixture_id=%s AND customer_id=%s AND transaction_type='receipt'
            ORDER BY transaction_date DESC, id DESC
            LIMIT 1
            """,
            (fixture_id, customer_id)
        )
        last_return = db.execute_query(
            """
            SELECT id, transaction_date, order_no, operator, note
            FROM material_transactions
            WHERE fixture_id=%s AND customer_id=%s AND transaction_type='return'
            ORDER BY transaction_date DESC, id DESC
            LIMIT 1
            """,
            (fixture_id, customer_id)
        )

        usage_logs = db.execute_query(
            """
            SELECT id, used_at, station_id, operator, note
            FROM usage_logs
            WHERE fixture_id=%s AND customer_id=%s
            ORDER BY used_at DESC
            LIMIT 100
            """,
            (fixture_id, customer_id)
        )

        replacement_logs = db.execute_query(
            """
            SELECT *
            FROM replacement_logs
            WHERE fixture_id=%s AND customer_id=%s
            ORDER BY replacement_date DESC
            LIMIT 100
            """,
            (fixture_id, customer_id)
        )

        serials = db.execute_query(
            """
            SELECT *
            FROM fixture_serials
            WHERE fixture_id=%s AND customer_id=%s
            ORDER BY serial_number
            """,
            (fixture_id, customer_id)
        )

        return {
            "fixture": fixture_rows[0],
            "serials": serials,
            "last_receipt": last_receipt[0] if last_receipt else None,
            "last_return": last_return[0] if last_return else None,
            "usage_logs": usage_logs,
            "replacement_logs": replacement_logs,
        }

    except HTTPException:
        raise
    except Exception:
        print("❌ [get_fixture_detail]\n", traceback.format_exc())
        raise HTTPException(status_code=500, detail="查詢治具詳細資料失敗")



# ============================================================
# 查詢治具列表 (LIST)
# ============================================================
@router.get("", summary="查詢治具列表")
async def list_fixtures(
    skip: int = Query(0, ge=0),
    limit: int = Query(8, ge=1, le=200),
    search: Optional[str] = Query(None),
    owner_id: Optional[int] = Query(None),
    storage: Optional[str] = Query(None),



    user=Depends(get_current_user),
    customer_id: str = Depends(get_current_customer_id),
):
    try:
        where = ["f.customer_id = %s"]
        where.append("f.is_scrapped = 0")
        params = [customer_id]

        if search:
            where.append("(f.id LIKE %s OR f.fixture_name LIKE %s)")
            like = f"%{search}%"
            params.extend([like, like])

        if storage:
            where.append("f.storage_location = %s")
            params.append(storage)

        if owner_id:
            where.append("f.owner_id = %s")
            params.append(owner_id)

        where_sql = " AND ".join(where)

        total = db.execute_query(
            f"SELECT COUNT(*) AS total FROM fixtures f WHERE {where_sql}",
            tuple(params)
        )[0]["total"]

        rows = db.execute_query(
            f"""
            SELECT
                f.*,
                u.username AS owner_name
            FROM fixtures f
            LEFT JOIN owners o ON f.owner_id = o.id
            LEFT JOIN users u ON u.id = o.primary_user_id
            WHERE {where_sql}
            ORDER BY f.id
            LIMIT %s OFFSET %s
            """,
            tuple(params + [limit, skip])
        )

        return {
            "total": total,
            "fixtures": rows
        }

    except Exception:
        print("❌ [list_fixtures]\n", traceback.format_exc())
        raise HTTPException(500, "查詢治具列表失敗")



# ============================================================
# 更新治具 (UPDATE) - 動態路徑
# ============================================================
@router.put("/{fixture_id}", response_model=FixtureResponse, summary="更新治具")
async def update_fixture(
    fixture_id: str,
    data: FixtureUpdate,
    user=Depends(get_current_user),
    customer_id: str = Depends(get_current_customer_id),
):
    try:
        # -------------------------------------------------
        # 1️⃣ 確認治具存在
        # -------------------------------------------------
        exists = db.execute_query(
            """
            SELECT id
            FROM fixtures
            WHERE id = %s AND customer_id = %s
            """,
            (fixture_id, customer_id)
        )
        if not exists:
            raise HTTPException(status_code=404, detail="治具不存在")

        # -------------------------------------------------
        # 2️⃣ 組 UPDATE 欄位
        # -------------------------------------------------
        update_fields = []
        params = []

        for field, value in data.dict(exclude_unset=True).items():
            if field == "status":
                continue

            if value is None:
                continue

            # Enum 安全處理
            if hasattr(value, "value"):
                value = value.value

            update_fields.append(f"{field} = %s")
            params.append(value)

        if not update_fields:
            raise HTTPException(status_code=400, detail="沒有要更新的欄位")

        params.extend([fixture_id, customer_id])

        update_sql = f"""
            UPDATE fixtures
            SET {', '.join(update_fields)}
            WHERE id = %s AND customer_id = %s
        """

        db.execute_update(update_sql, tuple(params))

        # -------------------------------------------------
        # 3️⃣ 回傳更新後資料（走標準 get_fixture）
        # -------------------------------------------------
        return await get_fixture(
            fixture_id=fixture_id,
            user=user,
            customer_id=customer_id,
        )

    except HTTPException:
        raise
    except Exception:
        print("❌ [update_fixture]\n", traceback.format_exc())
        raise HTTPException(status_code=500, detail="更新治具失敗")



@router.post("/{fixture_id}/scrap", summary="報廢治具")
async def scrap_fixture(
    fixture_id: str,
    admin=Depends(get_current_admin),
    customer_id: str = Depends(get_current_customer_id),
):
    try:
        # 1️⃣ 先讀治具狀態（給前端用）
        rows = db.execute_query(
            """
            SELECT
                in_stock_qty,
                is_scrapped
            FROM fixtures
            WHERE id = %s AND customer_id = %s
            """,
            (fixture_id, customer_id)
        )

        if not rows:
            raise HTTPException(status_code=404, detail="治具不存在")

        if rows[0]["is_scrapped"] == 1:
            raise HTTPException(status_code=400, detail="治具已是報廢狀態")

        # 2️⃣ 執行報廢（不動歷史資料）
        db.execute_update(
            """
            UPDATE fixtures
            SET
                is_scrapped = 1,
                scrapped_at = NOW(),
                updated_at = NOW()
            WHERE id = %s AND customer_id = %s
            """,
            (fixture_id, customer_id)
        )

        return {
            "message": "治具已報廢",
            "fixture_id": fixture_id
        }

    except HTTPException:
        raise
    except Exception:
        print("❌ [scrap_fixture]\n", traceback.format_exc())
        raise HTTPException(status_code=500, detail="報廢治具失敗")




@router.post("/{fixture_id}/restore", summary="還原報廢治具")
async def restore_fixture(
    fixture_id: str,
    admin=Depends(get_current_admin),
    customer_id: str = Depends(get_current_customer_id),
):
    try:
        affected = db.execute_update(
            """
            UPDATE fixtures
            SET
                is_scrapped = 0,
                scrapped_at = NULL,
                updated_at = NOW()
            WHERE id = %s
              AND customer_id = %s
              AND is_scrapped = 1
            """,
            (fixture_id, customer_id)
        )

        if affected == 0:
            raise HTTPException(status_code=400, detail="治具未處於報廢狀態")

        return {
            "message": "治具已還原",
            "fixture_id": fixture_id
        }

    except HTTPException:
        raise
    except Exception:
        print("❌ [restore_fixture]\n", traceback.format_exc())
        raise HTTPException(status_code=500, detail="還原治具失敗")

@router.get("/{fixture_id}/scrap-info", summary="取得治具報廢確認資訊")
async def get_fixture_scrap_info(
    fixture_id: str,
    admin=Depends(get_current_admin),
    customer_id: str = Depends(get_current_customer_id),
):
    try:
        rows = db.execute_query(
            """
            SELECT
                in_stock_qty,
                is_scrapped
            FROM fixtures
            WHERE id = %s AND customer_id = %s
            """,
            (fixture_id, customer_id)
        )

        if not rows:
            raise HTTPException(status_code=404, detail="治具不存在")

        if rows[0]["is_scrapped"] == 1:
            raise HTTPException(status_code=400, detail="治具已是報廢狀態")

        tx_exists = db.execute_query(
            """
            SELECT 1
            FROM material_transactions
            WHERE fixture_id = %s AND customer_id = %s
            LIMIT 1
            """,
            (fixture_id, customer_id)
        )

        return {
            "fixture_id": fixture_id,
            "in_stock_qty": rows[0]["in_stock_qty"] or 0,
            "has_transactions": bool(tx_exists),
        }

    except HTTPException:
        raise
    except Exception:
        print("❌ [get_fixture_scrap_info]\n", traceback.format_exc())
        raise HTTPException(status_code=500, detail="取得報廢確認資訊失敗")
