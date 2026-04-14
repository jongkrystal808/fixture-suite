"""
治具 Excel 匯入 Router（v6）
- 路徑：POST /fixtures/import
- Header-based customer context
- 僅允許「新增治具」
- fixture_id 已存在 → 視為錯誤
- 任一列錯誤 → 全部 rollback
- 匯入回報格式完全對齊 models import
"""

from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
from openpyxl import load_workbook
from io import BytesIO

from backend.app.database import db
from backend.app.dependencies import (
    get_current_user,
    get_current_customer_id,
)

router = APIRouter(prefix="/fixtures", tags=["Fixtures Import"])


# ============================================================
# Utils
# ============================================================

def compress_rows(rows: list[int]) -> list[str]:
    """
    將 [2,3,4,7,9] 壓縮為 ['2-4', '7', '9']
    """
    if not rows:
        return []

    rows = sorted(set(rows))
    result = []

    start = prev = rows[0]
    for r in rows[1:]:
        if r == prev + 1:
            prev = r
            continue
        result.append(f"{start}-{prev}" if start != prev else f"{start}")
        start = prev = r

    result.append(f"{start}-{prev}" if start != prev else f"{start}")
    return result


# ============================================================
# Router
# ============================================================

@router.post("/import", summary="匯入治具（XLSX）")
async def import_fixtures_xlsx(
    file: UploadFile = File(...),
    user=Depends(get_current_user),
    customer_id: str = Depends(get_current_customer_id),
):

    # =================================================
    # 1️⃣ 基本檢查
    # =================================================
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="只接受 .xlsx 檔案")

    content = await file.read()
    wb = load_workbook(BytesIO(content))

    if "fixtures" not in wb.sheetnames:
        raise HTTPException(
            status_code=400,
            detail='XLSX 需包含 "fixtures" 工作表'
        )

    ws = wb["fixtures"]

    # =================================================
    # 2️⃣ Header 檢查
    # =================================================
    header = [str(c.value).strip() if c.value else "" for c in ws[1]]

    REQUIRED_COLS = ["id", "fixture_name"]

    for col in REQUIRED_COLS:
        if col not in header:
            raise HTTPException(
                status_code=400,
                detail=f"缺少必要欄位：{col}"
            )

    idx = {name: header.index(name) for name in header}

    # =================================================
    # 3️⃣ 初始化回報
    # =================================================
    result = {
        "fixtures": {
            "imported": 0,
            "skipped": 0,
            "success_rows": [],
            "skipped_rows": [],
        },
        "errors": [],
    }

    # 記錄成功列的 row_no → fixture_id，供第 6️⃣ 步 cache 重建使用
    imported_fixtures: dict[int, str] = {}

    conn = db.get_conn()
    cursor = conn.cursor()

    try:
        # =================================================
        # 4️⃣ 逐行處理
        # =================================================
        # 使用 values_only=False，從 cell 物件取得 Excel 真實行號
        for row in ws.iter_rows(min_row=2, values_only=False):
            row_no = row[0].row  # ← 直接取 Excel 真實行號，不依賴 enumerate
            row_values = tuple(cell.value for cell in row)

            try:
                # 全部欄位皆為 None → 真正的空行
                if not any(row_values):
                    result["fixtures"]["skipped"] += 1
                    result["fixtures"]["skipped_rows"].append(row_no)
                    continue

                fixture_id = str(row_values[idx["id"]] or "").strip()
                fixture_name = str(row_values[idx["fixture_name"]] or "").strip()

                if not fixture_id or not fixture_name:
                    result["fixtures"]["skipped"] += 1
                    result["fixtures"]["skipped_rows"].append(row_no)
                    continue

                fixture_type = (
                    str(row_values[idx["fixture_type"]]).strip()
                    if "fixture_type" in idx and row_values[idx["fixture_type"]] is not None
                    else ""
                )

                storage_location = (
                    str(row_values[idx["storage_location"]]).strip()
                    if "storage_location" in idx and row_values[idx["storage_location"]] is not None
                    else ""
                )

                replacement_cycle = (
                    int(row_values[idx["replacement_cycle"]])
                    if "replacement_cycle" in idx and row_values[idx["replacement_cycle"]] is not None
                    else None
                )

                cycle_unit_raw = (
                    str(row_values[idx["cycle_unit"]]).strip().lower()
                    if "cycle_unit" in idx and row_values[idx["cycle_unit"]] is not None
                    else "uses"
                )

                note = (
                    str(row_values[idx["note"]]).strip()
                    if "note" in idx and row_values[idx["note"]] is not None
                    else ""
                )

                # ---------- 驗證 / 正規化 ----------
                if cycle_unit_raw == "none":
                    cycle_unit = None
                    replacement_cycle = None
                elif cycle_unit_raw in ("uses", "days"):
                    cycle_unit = cycle_unit_raw
                else:
                    raise ValueError("cycle_unit 必須是 uses / days / none")

                # ---------- 不允許已存在 ----------
                cursor.execute(
                    """
                    SELECT 1
                    FROM fixtures
                    WHERE customer_id=%s AND id=%s
                    """,
                    (customer_id, fixture_id)
                )
                if cursor.fetchone():
                    raise ValueError(
                        f"治具編號 {fixture_id} 已存在，禁止重複匯入"
                    )

                # ---------- INSERT（v6 cache 初始化） ----------
                cursor.execute(
                    """
                    INSERT INTO fixtures (
                        customer_id,
                        id,
                        fixture_name,
                        fixture_type,

                        self_purchased_qty,
                        customer_supplied_qty,
                        in_stock_qty,
                        deployed_qty,
                        maintenance_qty,
                        scrapped_qty,
                        returned_qty,

                        storage_location,
                        replacement_cycle,
                        cycle_unit,
                        note
                    )
                    VALUES (
                        %s,%s,%s,%s,
                        0,0,0,0,0,0,0,
                        %s,%s,%s,%s
                    )
                    """,
                    (
                        customer_id,
                        fixture_id,
                        fixture_name,
                        fixture_type,
                        storage_location,
                        replacement_cycle,
                        cycle_unit,
                        note,
                    )
                )

                result["fixtures"]["imported"] += 1
                result["fixtures"]["success_rows"].append(row_no)
                imported_fixtures[row_no] = fixture_id  # ← 記錄供 cache 重建用

            except Exception as e:
                result["errors"].append(f"第 {row_no} 行：{str(e)}")

        # =================================================
        # 5️⃣ 錯誤處理
        # =================================================
        if result["errors"]:
            conn.rollback()
            raise HTTPException(
                status_code=400,
                detail={
                    "message": "匯入失敗，資料未寫入",
                    **result,
                }
            )

        if result["fixtures"]["imported"] == 0:
            conn.rollback()
            return {
                "message": "未匯入任何有效治具資料",
                "fixtures": {
                    "imported": 0,
                    "skipped": result["fixtures"]["skipped"],
                    "success_rows": [],
                    "skipped_rows": compress_rows(
                        result["fixtures"]["skipped_rows"]
                    ),
                },
                "errors": [],
                "warning": "檔案中沒有可匯入的有效資料列",
            }

        # =================================================
        # 6️⃣ v6：重建 cache
        # 直接用 imported_fixtures dict，不再重新讀 worksheet
        # =================================================
        for fixture_id in imported_fixtures.values():
            cursor.execute(
                "CALL sp_rebuild_fixture_quantities_v6(%s, %s)",
                (customer_id, fixture_id)
            )

        conn.commit()

    finally:
        cursor.close()
        conn.close()

    # =================================================
    # 7️⃣ 成功回傳
    # =================================================
    return {
        "message": "匯入完成",
        "fixtures": {
            "imported": result["fixtures"]["imported"],
            "skipped": result["fixtures"]["skipped"],
            "success_rows": compress_rows(
                result["fixtures"]["success_rows"]
            ),
            "skipped_rows": compress_rows(
                result["fixtures"]["skipped_rows"]
            ),
        },
        "errors": [],
    }